"""主程序入口 - 美国电话号码反查"""
import logging
import csv
import sys
from pathlib import Path
from datetime import datetime

try:
    import openpyxl
    from openpyxl import load_workbook
except ImportError:
    print("❌ 缺少 openpyxl 库，请运行: pip install openpyxl")
    sys.exit(1)

import config
from api_query import WhitepagesAPI

# 配置日志
logging.basicConfig(
    level=getattr(logging, config.LOG_LEVEL),
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(config.LOG_FILE, encoding='utf-8'),
        logging.StreamHandler()
    ]
)

logger = logging.getLogger(__name__)


def read_input_file(filename: str) -> list:
    """读取输入 XLSX 文件
    
    Args:
        filename: XLSX 文件路径
        
    Returns:
        电话号码列表
    """
    try:
        if not Path(filename).exists():
            logger.error(f"❌ 输入文件不存在: {filename}")
            return []
        
        logger.info(f"正在读取输入文件: {filename}")
        
        # 使用当日日期作为列名
        column_name = config.TODAY_DATE
        
        workbook = load_workbook(filename)
        worksheet = workbook.active
        
        phones = []
        header_found = False
        col_index = None
        
        # 查找列
        for row in worksheet.iter_rows(min_row=1, max_row=1, values_only=True):
            for idx, cell_value in enumerate(row, 1):
                if cell_value == column_name:
                    header_found = True
                    col_index = idx
                    logger.info(f"✓ 找到列: {column_name} (第 {idx} 列)")
                    break
        
        if not header_found:
            logger.warning(f"⚠️ 未找到列: {column_name}，使用第一列")
            col_index = 1
        
        # 读取数据
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            if row[col_index - 1]:
                phone = str(row[col_index - 1]).strip()
                if phone and phone.lower() != "nan":
                    phones.append(phone)
        
        logger.info(f"✓ 读取成功，共 {len(phones)} 条记录")
        return phones
        
    except Exception as e:
        logger.error(f"❌ 读取输入文件失败: {e}")
        return []


def save_results(results: list, filename: str) -> bool:
    """保存查询结果到 CSV 文件
    
    Args:
        results: 查询结果列表
        filename: 输出文件路径
        
    Returns:
        是否保存成功
    """
    try:
        # 创建输出目录
        Path(filename).parent.mkdir(parents=True, exist_ok=True)
        
        if not results:
            logger.warning("⚠️ 无结果可保存")
            return False
        
        logger.info(f"正在保存结果到: {filename}")
        
        # 获取字段名
        fieldnames = ["phone", "name", "age_range", "property_value", "occupation", "status"]
        
        # 写入 CSV
        with open(filename, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            
            for result in results:
                # 确保所有字段都存在
                row = {field: result.get(field, "") for field in fieldnames}
                writer.writerow(row)
        
        logger.info(f"✓ 保存成功，共 {len(results)} 条记录到 {filename}")
        return True
        
    except Exception as e:
        logger.error(f"❌ 保存结果失败: {e}")
        return False


def main():
    """主程序"""
    logger.info("=" * 70)
    logger.info("🚀 美国电话号码反查系统 - Whitepages Pro API")
    logger.info("=" * 70)
    
    try:
        # 1. 读取输入文件
        logger.info("\n📖 第一步: 读取输入文件")
        phones = read_input_file(config.INPUT_FILE)
        
        if not phones:
            logger.error("❌ 没有获取到电话号码，程序退出")
            return False
        
        # 2. 初始化 API
        logger.info("\n🔌 第二步: 初始化 Whitepages Pro API")
        api = WhitepagesAPI()
        
        # 3. 批量查询
        logger.info("\n🔍 第三步: 批量查询电话号码")
        results = api.batch_query(phones)
        
        # 4. 保存结果
        logger.info("\n💾 第四步: 保存结果到 CSV")
        success = save_results(results, config.OUTPUT_FILE)
        
        # 5. 关闭 API
        api.close()
        
        # 输出统计信息
        logger.info("\n" + "=" * 70)
        logger.info("📊 查询统计")
        logger.info("=" * 70)
        logger.info(f"总查询数: {len(phones)}")
        logger.info(f"成功数: {len([r for r in results if r.get('status') == '成功'])}")
        logger.info(f"失败数: {len([r for r in results if r.get('status') != '成功'])}")
        logger.info(f"输出文件: {config.OUTPUT_FILE}")
        logger.info("=" * 70)
        
        return success
        
    except KeyboardInterrupt:
        logger.warning("⚠️ 用户中断程序")
        return False
    except Exception as e:
        logger.error(f"❌ 程序异常: {e}")
        import traceback
        logger.error(traceback.format_exc())
        return False


if __name__ == "__main__":
    success = main()
    sys.exit(0 if success else 1)
