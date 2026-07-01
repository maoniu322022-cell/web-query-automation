import logging
from typing import List, Dict
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from threading import Lock
import os

logger = logging.getLogger(__name__)

# 全局锁，用于线程安全的结果保存
results_lock = Lock()

class DataHandler:
    def __init__(self, input_file="names.txt", output_file="search_results.xlsx"):
        self.input_file = input_file
        self.output_file = output_file
    
    def load_names(self, filename: str) -> List[str]:
        """读取名字列表"""
        try:
            with open(filename, 'r', encoding='utf-8') as f:
                names = [line.strip() for line in f if line.strip()]
            return names
        except FileNotFoundError:
            logger.error(f"File not found: {filename}")
            return []
        except Exception as e:
            logger.error(f"Error reading file: {e}")
            return []
    
    def save_results(self, results: List[Dict], filename: str):
        """保存结果到 Excel"""
        try:
            # 如果文件存在，加载现有数据
            if os.path.exists(filename):
                wb = load_workbook(filename)
                ws = wb.active
                start_row = ws.max_row + 1
            else:
                wb = Workbook()
                ws = wb.active
                ws.title = "Results"
                start_row = 1
                
                # 设置表头
                headers = ["Name", "Age", "Phone", "Location"]
                ws.append(headers)
                
                # 设置表头样式
                header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF")
                
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                
                # 调整列宽
                ws.column_dimensions['A'].width = 20
                ws.column_dimensions['B'].width = 10
                ws.column_dimensions['C'].width = 18
                ws.column_dimensions['D'].width = 25
                
                start_row = 2
            
            # 添加新数据
            for result in results:
                ws.append([
                    result.get("name", ""),
                    result.get("age", ""),
                    result.get("phone", ""),
                    result.get("location", "")
                ])
            
            wb.save(filename)
            logger.info(f"✓ 结果已保存到 {filename} (共 {len(results)} 条新记录)")
            
        except Exception as e:
            logger.error(f"Error saving results: {e}")
